# Importing Modu
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# EXL_VIDEO_TITLE = 1
# EXL_ID_COL = 2
# EXL_ENGLISH_COL = 3
# EXL_START_COL = 4
# EXL_END_COL = 5
# EXL_TRANSLATION_COL = 6

# EXL_VIDEO_TITLE = 1
# EXL_ID_COL = 2
# EXL_ENGLISH_COL = 6
# EXL_START_COL = 4
# EXL_END_COL = 5
# EXL_TRANSLATION_COL = 5

EXL_VIDEO_TITLE = 1
EXL_ID_COL = 2
EXL_ENGLISH_COL = 3
EXL_START_COL = 4
EXL_END_COL = 5
EXL_TRANSLATION_COL = 6


# EXCEL_FILE_TITLE = 'ARABIC_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'ALBANIAN_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'DARI_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'FRENCH_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'PASHTO_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'SORANI_KURDISH_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'TIGRINYA_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'UKRAINIAN_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'VIETNAMESE_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'ROMANIAN_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'AMHARIC_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'FARSI_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'SPANISH_SUBTITLES_PYE2.xlsx'
# EXCEL_FILE_TITLE = 'Amharic PYE2 Subtitles TRANSLATED AMHARIC.xlsx'

EXCEL_FILE_TITLE = 'PYE_SUBTITLES_MASTER_SHEET.xlsx'


FOLDER_NAME = str.removesuffix(EXCEL_FILE_TITLE, '.xlsx')
ROOT_PATH = "/Users/paullunn/Documents/GitHub/subtitler/"
OUTPUT_FOLDER_PATH = ROOT_PATH + FOLDER_NAME
SUBTITLE_FILE_SUFFIX = ".srt"

class FilmData:
    name = ""
    start_row = 0
    end_row = 0

    # constructor function    
    def __init__(self, name, start, end):
        self.name = name
        self.start_row = start
        self.end_row = end
    
    def __str__(self):
        return "{} start = {} end = {}".format(self.name, self.start_row, self.end_row)

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def getExcelRowData(excel_sheet, excel_row):
    id = excel_sheet.cell(row=excel_row, column=EXL_ID_COL).value
    start_time = excel_sheet.cell(row=excel_row, column=EXL_START_COL).value
    end_time = excel_sheet.cell(row=excel_row, column=EXL_END_COL).value
    text = excel_sheet.cell(row=excel_row, column=EXL_TRANSLATION_COL).value
    # english = excel_sheet.cell(row=excel_row, column=EXL_ENGLISH_COL).value
    english = "None"

    return int(id), start_time, end_time, text, english

def create_srt_entry(id, start_time, end_time , text):
    # 3 00:01:23,862 --> 00:01:25,091 Morning Post.
    subtitle = '{} {} --> {} {}'.format(id, start_time, end_time, text)
    subtitle = subtitle + "\n"
    return subtitle

def find_all_videos(excel_sheet):
    videos = [] 
    for cell in excel_sheet['A']:
        name = cell.value
        if name and name != "Video Title":
            # print(name, cell.row)
            filmdata = FilmData(name, cell.row, 0)
            videos.append (filmdata)
    
    num_rows = excel_sheet.max_row
    for row in range (0, len(videos)-1):
        # print(videos[row])
        next_start = videos[row+1].start_row
        videos[row].end_row = next_start -1
        # print(videos[row])

    videos[len(videos)-1].end_row = num_rows
    return videos

def print_all_videos(videos):
    for video in videos:
        print(video)

            


workbook = load_workbook(filename=EXCEL_FILE_TITLE)
sheet = workbook.active
max_rows_in_spreadsheet = get_maximum_rows(sheet_object=sheet)
# print("******** ",max_rows_in_spreadsheet)

videos = find_all_videos(sheet)
num_videos = len(videos)
print("Workbook = ",EXCEL_FILE_TITLE)
print("num_videos = {}".format(num_videos))
# print_all_videos(videos)

if not os.path.exists(OUTPUT_FOLDER_PATH):
    os.mkdir(OUTPUT_FOLDER_PATH)

for video_index, video in enumerate(videos):
    subtitle_file = open(OUTPUT_FOLDER_PATH + "/" + video.name + SUBTITLE_FILE_SUFFIX, "w")
    
    if video_index == num_videos-1:
        video.end_row = max_rows_in_spreadsheet+1
    print(" index = ", video_index, " start = ", video.start_row, " end = ", video.end_row, " titles = ",video.name)
    
    for row in range (video.start_row, video.end_row):
        id, start, end, translation, english = getExcelRowData(sheet, row)
        subtitle_file.write(str(id) + "\n")
        timing =  '{} --> {}'.format(start, end)
        subtitle_file.write(timing + "\n")
        subtitle_file.write(translation + "\n" + "\n")


    subtitle_file.close()

